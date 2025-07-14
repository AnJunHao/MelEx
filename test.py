import exmel
from pathlib import Path
from openpyxl import load_workbook
from typing import NamedTuple
import numpy as np
import pandas as pd
import time
from dataclasses import dataclass

@dataclass
class AlignParams:
    split_melody: bool = True
    hop_length: int = 2

params = AlignParams()

base_path = Path("mid_files/test")
xlsx_file = base_path / "table.xlsx"
workbook = load_workbook(xlsx_file)
sheet = workbook["Sheet1"]

class Row(NamedTuple):
    id_: str
    hash_: str
    song_name: str
    singer: str

table: list[Row] = []
for row in sheet.iter_rows(min_row=2, max_col=10, values_only=True):
    if row[0] is None:
        break
    table.append(Row(*[str(i) for i in row[:4]])) # type: ignore

# Store results for statistical analysis
baseline_results = []
alignment_results = []
sample_names = []
alignment_runtimes = []

result_dir = Path("results")
result_dir.mkdir(exist_ok=True)

for row in table:
    print(f"Aligning {row.song_name}...")

    melody = base_path / "melody" / f"{row.hash_}.est.note"
    transcription = base_path / "piano_melody_test_samples" / row.id_ / f"{row.id_}_transcription.mid"
    ground_truth = exmel.Melody(
        base_path / "piano_melody_test_samples" / row.id_ / f"{row.id_}_rectified.mid",
        track_idx=3)
    baseline = exmel.Melody(
        base_path / "piano_melody_test_samples" / row.id_ / f"{row.id_}.mid",
        track_idx=3)

    # Time the alignment operation
    start_time = time.time()
    alignment = exmel.align(melody, transcription,
        exmel.duration_adjusted_weighted_sum_velocity,
        split_melody=params.split_melody,
        hop_length=params.hop_length)
    end_time = time.time()
    runtime = end_time - start_time

    result_baseline = exmel.evaluate_melody(ground_truth, baseline, plot=False)
    result_alignment = exmel.evaluate_melody(
        ground_truth, alignment.events,
        plot=False, save_path=f"results/{row.song_name}_eval.png")

    exmel.save_melody(alignment.events, f"results/{row.song_name}.mid")
    exmel.plot_alignment(alignment, melody, transcription, ground_truth, f"results/{row.song_name}_align.png")

    # Print results
    print(f"Baseline: {result_baseline.precision*100:.1f}% {result_baseline.recall*100:.1f}% {result_baseline.f1_score*100:.1f}%")
    print(f"Alignment: {result_alignment.precision*100:.1f}% {result_alignment.recall*100:.1f}% {result_alignment.f1_score*100:.1f}%")
    print(f"Runtime: {runtime:.3f}s")
    
    # Store results
    baseline_results.append(result_baseline)
    alignment_results.append(result_alignment)
    sample_names.append(row.song_name)
    alignment_runtimes.append(runtime)

# Calculate aggregate statistics
def calculate_stats(results):
    precisions = [r.precision for r in results]
    recalls = [r.recall for r in results]
    f1_scores = [r.f1_score for r in results]
    
    return {
        'precision': {
            'mean': np.mean(precisions),
            'std': np.std(precisions),
            'min': np.min(precisions),
            'max': np.max(precisions)
        },
        'recall': {
            'mean': np.mean(recalls),
            'std': np.std(recalls),
            'min': np.min(recalls),
            'max': np.max(recalls)
        },
        'f1_score': {
            'mean': np.mean(f1_scores),
            'std': np.std(f1_scores),
            'min': np.min(f1_scores),
            'max': np.max(f1_scores)
        },
        'tp_total': sum(r.tp for r in results),
        'fp_total': sum(r.fp for r in results),
        'fn_total': sum(r.fn for r in results)
    }

baseline_stats = calculate_stats(baseline_results)
alignment_stats = calculate_stats(alignment_results)

# Print detailed statistics
print("\n" + "="*80)
print("EVALUATION RESULTS SUMMARY")
print("="*80)

print(f"\nTested on {len(table)} samples")

print("\nBASELINE RESULTS:")
print("-" * 40)
print(f"Precision: {baseline_stats['precision']['mean']:.3f} ± {baseline_stats['precision']['std']:.3f}")
print(f"Recall:    {baseline_stats['recall']['mean']:.3f} ± {baseline_stats['recall']['std']:.3f}")
print(f"F1 Score:  {baseline_stats['f1_score']['mean']:.3f} ± {baseline_stats['f1_score']['std']:.3f}")
print(f"Total TP: {baseline_stats['tp_total']}, FP: {baseline_stats['fp_total']}, FN: {baseline_stats['fn_total']}")

print("\nALIGNMENT RESULTS:")
print("-" * 40)
print(f"Precision: {alignment_stats['precision']['mean']:.3f} ± {alignment_stats['precision']['std']:.3f}")
print(f"Recall:    {alignment_stats['recall']['mean']:.3f} ± {alignment_stats['recall']['std']:.3f}")
print(f"F1 Score:  {alignment_stats['f1_score']['mean']:.3f} ± {alignment_stats['f1_score']['std']:.3f}")
print(f"Total TP: {alignment_stats['tp_total']}, FP: {alignment_stats['fp_total']}, FN: {alignment_stats['fn_total']}")

print("\nIMPROVEMENT:")
print("-" * 40)
precision_improvement = alignment_stats['precision']['mean'] - baseline_stats['precision']['mean']
recall_improvement = alignment_stats['recall']['mean'] - baseline_stats['recall']['mean']
f1_improvement = alignment_stats['f1_score']['mean'] - baseline_stats['f1_score']['mean']

print(f"Precision: {precision_improvement:+.3f} ({precision_improvement/baseline_stats['precision']['mean']*100:+.1f}%)")
print(f"Recall:    {recall_improvement:+.3f} ({recall_improvement/baseline_stats['recall']['mean']*100:+.1f}%)")
print(f"F1 Score:  {f1_improvement:+.3f} ({f1_improvement/baseline_stats['f1_score']['mean']*100:+.1f}%)")

print("\nALIGNMENT RUNTIME:")
print("-" * 40)
print(f"Average: {np.mean(alignment_runtimes):.3f}s")
print(f"Std Dev: {np.std(alignment_runtimes):.3f}s")
print(f"Min:     {np.min(alignment_runtimes):.3f}s")
print(f"Max:     {np.max(alignment_runtimes):.3f}s")
print(f"Total:   {np.sum(alignment_runtimes):.3f}s")

# Create detailed per-sample results DataFrame
results_data = []
for i, (name, baseline, alignment, runtime) in enumerate(zip(sample_names, baseline_results, alignment_results, alignment_runtimes)):
    results_data.append({
        'Sample': name,
        'Baseline_Precision': baseline.precision,
        'Baseline_Recall': baseline.recall,
        'Baseline_F1': baseline.f1_score,
        'Alignment_Precision': alignment.precision,
        'Alignment_Recall': alignment.recall,
        'Alignment_F1': alignment.f1_score,
        'Precision_Improvement': alignment.precision - baseline.precision,
        'Recall_Improvement': alignment.recall - baseline.recall,
        'F1_Improvement': alignment.f1_score - baseline.f1_score,
        'Runtime_seconds': runtime
    })

# Add a final row with the aggregate statistics
results_data.append({
    'Sample': 'Aggregate',
    'Baseline_Precision': baseline_stats['precision']['mean'],
    'Baseline_Recall': baseline_stats['recall']['mean'],
    'Baseline_F1': baseline_stats['f1_score']['mean'],
    'Alignment_Precision': alignment_stats['precision']['mean'],
    'Alignment_Recall': alignment_stats['recall']['mean'],
    'Alignment_F1': alignment_stats['f1_score']['mean'],
    'Precision_Improvement': alignment_stats['precision']['mean'] - baseline_stats['precision']['mean'],
    'Recall_Improvement': alignment_stats['recall']['mean'] - baseline_stats['recall']['mean'],
    'F1_Improvement': alignment_stats['f1_score']['mean'] - baseline_stats['f1_score']['mean'],
    'Runtime_seconds': np.mean(alignment_runtimes)
})

df = pd.DataFrame(results_data)

# Save results to CSV
df.to_csv('evaluation_results.csv', index=False)
print(f"\nDetailed results saved to 'evaluation_results.csv'")