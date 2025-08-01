from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Define the statistics as global constants
PERCENTILE_STATS = {
    'alignment_over_performance': {'5%': 0.641018, '50%': 0.770082, '95%': 0.915470},
    'melody_over_performance': {'5%': 0.640966, '50%': 0.776179, '95%': 0.932307},
    'structural_recurrence': {'5%': 0.826956, '50%': 0.993821, '95%': 1.000000},
    'pred_f1': {'5%': 0.895976, '50%': 0.963794, '95%': 0.985737},
    'pred_precision': {'5%': 0.940084, '50%': 0.979746, '95%': 0.991077},
    'pred_recall': {'5%': 0.869881, '50%': 0.950738, '95%': 0.979879},
    'above_between': {'5%': 0.000000, '50%': 0.006674, '95%': 0.031620},
    'between': {'5%': 0.007953, '50%': 0.026763, '95%': 0.074561},
    'error': {'5%': 0.007732, '50%': 0.020500, '95%': 0.050531},
    'miss': {'5%': 0.002068, '50%': 0.010914, '95%': 0.036747},
    'shadow': {'5%': 0.000000, '50%': 0.006007, '95%': 0.048507}
}

# Define which metrics are "higher is better" vs "lower is better"
HIGHER_IS_BETTER = ['alignment_over_performance', 'melody_over_performance', 
                    'structural_recurrence', 'pred_f1', 'pred_precision', 'pred_recall']
LOWER_IS_BETTER = ['above_between', 'between', 'error', 'miss', 'shadow']

# Define the primary metric
PRIMARY_METRIC = 'pred_f1'

# Column descriptions in multiple languages
COLUMN_DESCRIPTIONS = {
    'name': {
        'en': 'Model or experiment identifier',
        'zh': '歌曲名或标识符'
    },
    'alignment_over_performance': {
        'en': 'Duration ratio of extracted melody to piano transcription duration, excluding non-melodic segments (interludes, etc.)',
        'zh': '提取得到的旋律的时长 除以 钢琴转录的时长。旋律的时长在计算时排除了不包含旋律的段落（如间奏等无旋律段落不计入时长）'
    },
    'melody_over_performance': {
        'en': 'Duration ratio of original melody to piano transcription duration, excluding non-melodic segments (interludes, etc.)',
        'zh': '原曲旋律的时长 除以 钢琴转录的时长。旋律的时长在计算时排除了不包含旋律的段落（如间奏等无旋律段落不计入时长）'
    },
    'structural_recurrence': {
        'en': 'Structural recurrence rate - measures how well the original musical structure is reproduced in the piano transcription, higher values indicate greater structural similarity',
        'zh': '结构复现率，指原曲曲式结构在钢琴转录中复现的程度，数值越高表示结构相似度越高'
    },
    'pred_f1': {
        'en': '★ PRIMARY METRIC ★ Predicted F1 score, harmonic mean of precision and recall',
        'zh': '★ 主要指标 ★ 预测的F1分数，精确率和召回率的调和平均值'
    },
    'pred_precision': {
        'en': 'Predicted Precision - ratio of correct extracted notes to total extracted notes',
        'zh': '预测的精确率 - 正确的提取音符数与总提取音符数的比率'
    },
    'pred_recall': {
        'en': 'Predicted Recall - ratio of correct extracted notes to total actual notes',
        'zh': '预测的召回率 - 正确的提取音符数与实际总音符数的比率'
    },
    'above_between': {
        'en': 'High interval rate - proportion of notes with higher pitch appearing between adjacent notes in the extracted melody',
        'zh': '高间隔率，指提取得到的旋律中每两个相邻的音符之间，出现音高高于这两个音符的其他音符的比例'
    },
    'between': {
        'en': 'Middle interval rate - proportion of notes with pitch between or equal to adjacent notes in the extracted melody',
        'zh': '中间隔率，指提取得到的旋律中每两个相邻的音符之间，出现音高介于这两个音符之间或等于某个音符的其他音符的比例'
    },
    'error': {
        'en': 'Error rate - temporal difference between extracted melody from piano transcription and original melody, divided by piano transcription duration',
        'zh': '误差率：从钢琴转录提取的旋律与原曲旋律之间的时间误差 除以 钢琴转录的时长'
    },
    'miss': {
        'en': 'Miss rate - proportion of notes present in original melody but missing in piano transcription',
        'zh': '漏检率，指钢琴转录中不存在，但在原曲旋律中存在的音符的比例'
    },
    'shadow': {
        'en': 'Shadow rate - proportion of extracted melody notes that are overshadowed by higher-pitched notes in piano transcription within ±0.1 seconds',
        'zh': '遮蔽率。对于提取得到的旋律中的每个音符，若在钢琴转录中存在与该音符+-0.1秒的音高更高的音符，则认为该音符被遮蔽，遮蔽率统计了被遮蔽的音符占总提取音符的比例'
    }
}

def get_color_for_value(value, column_name, percentiles):
    """
    Determine the color based on the value and percentile thresholds.
    Returns a PatternFill object with the appropriate color.
    """
    p5 = percentiles['5%']
    p50 = percentiles['50%']
    p95 = percentiles['95%']
    
    # Define colors
    dark_green = PatternFill(start_color='00AA00', end_color='00AA00', fill_type='solid')
    light_green = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    yellow = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    light_red = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    dark_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    
    if column_name in HIGHER_IS_BETTER:
        # Higher is better
        if value >= p95:
            return dark_green  # Top 5%
        elif value >= p50:
            return light_green  # Between 50% and 95%
        elif value >= p5:
            return yellow  # Between 5% and 50%
        else:
            return light_red  # Bottom 5%
    else:
        # Lower is better
        if value <= p5:
            return dark_green  # Top 5% (lowest values)
        elif value <= p50:
            return light_green  # Between 5% and 50%
        elif value <= p95:
            return yellow  # Between 50% and 95%
        else:
            return light_red  # Worst 5% (highest values)

def format_dataframe_to_excel(df, output_path, language='en'):
    """
    Format a dataframe and save it as an Excel file with color-coded cells based on percentiles.
    
    Parameters:
    -----------
    df : pandas.DataFrame
        The dataframe to format
    output_path : str
        The path where the Excel file will be saved
    language : str
        Language for descriptions ('en' for English, 'zh' for Chinese)
    """
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Formatted Data"
    
    # Write the dataframe to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Find the column index for pred_f1
    pred_f1_col_idx = None
    for idx, col_name in enumerate(df.columns):
        if col_name == PRIMARY_METRIC:
            pred_f1_col_idx = idx + 1  # +1 because Excel is 1-indexed
            break
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_border = Border(
        left=Side(style='thick', color='000080'),  # Navy blue
        right=Side(style='thick', color='000080'),
        top=Side(style='thick', color='000080'),
        bottom=Side(style='thick', color='000080')
    )
    
    # Format headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    primary_header_fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')  # Navy blue for primary metric
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col_idx, cell in enumerate(ws[1]):
        cell.font = header_font
        cell.alignment = header_alignment
        
        # Special formatting for primary metric header
        if col_idx + 1 == pred_f1_col_idx:
            cell.fill = primary_header_fill
            cell.font = Font(bold=True, color='FFFF00', size=12)  # Yellow text, slightly larger
            cell.border = thick_border
        else:
            cell.fill = header_fill
            cell.border = thin_border
    
    # Apply color formatting to data cells
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        for col_idx, cell in enumerate(row):
            column_name = df.columns[col_idx]
            
            # Only apply color formatting to metric columns (not 'name')
            if column_name in PERCENTILE_STATS and cell.value is not None:
                try:
                    value = float(cell.value) # type: ignore
                    color_fill = get_color_for_value(value, column_name, PERCENTILE_STATS[column_name])
                    cell.fill = color_fill
                except (ValueError, TypeError):
                    # Skip if value cannot be converted to float
                    pass
            
            # Center align all cells
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply borders
            if col_idx + 1 == pred_f1_col_idx:
                cell.border = thick_border
                # Add subtle background tint to primary metric column
                if column_name != PRIMARY_METRIC or cell.value is None:
                    # If it's the name column or empty, just add light blue background
                    if not hasattr(cell, 'fill') or cell.fill.start_color.rgb == '00000000':
                        cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')  # Lavender
            else:
                cell.border = thin_border
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter # type: ignore
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Make pred_f1 column slightly wider
    if pred_f1_col_idx:
        pred_f1_letter = ws.cell(row=1, column=pred_f1_col_idx).column_letter # type: ignore
        ws.column_dimensions[pred_f1_letter].width = ws.column_dimensions[pred_f1_letter].width + 3
    
    # Add a legend sheet
    legend_ws = wb.create_sheet("Legend")
    
    # Build legend data
    legend_data = [
        ["Color Legend & Metric Information", ""],
        ["", ""],
        ["★ Primary Metric: pred_f1 (highlighted with thick blue border) ★", ""],
        ["", ""],
        ["Color Interpretation:", ""],
        ["", ""],
        ["For Higher-is-Better Metrics:", ""],
        ["Dark Green", "≥ 95th percentile (Top 5%)"],
        ["Light Green", "≥ 50th percentile"],
        ["Yellow", "≥ 5th percentile"],
        ["Light Red", "< 5th percentile (Bottom 5%)"],
        ["", ""],
        ["For Lower-is-Better Metrics:", ""],
        ["Dark Green", "≤ 5th percentile (Top 5%)"],
        ["Light Green", "≤ 50th percentile"],
        ["Yellow", "≤ 95th percentile"],
        ["Light Red", "> 95th percentile (Worst 5%)"],
        ["", ""],
        ["", ""],
        ["Metric Categories:", ""],
        ["", ""],
        ["Higher is Better (↑):", ""],
    ]
    
    # Add higher-is-better metrics
    for metric in HIGHER_IS_BETTER:
        if metric == PRIMARY_METRIC:
            legend_data.append(["", f"• {metric} ★ (PRIMARY METRIC)"])
        else:
            legend_data.append(["", f"• {metric}"])
    
    legend_data.extend([
        ["", ""],
        ["Lower is Better (↓):", ""]
    ])
    
    # Add lower-is-better metrics
    for metric in LOWER_IS_BETTER:
        legend_data.append(["", f"• {metric}"])
    
    # Write legend data
    for row_idx, row_data in enumerate(legend_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            legend_ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Apply colors to legend color samples
    legend_ws.cell(row=8, column=1).fill = PatternFill(start_color='00AA00', end_color='00AA00', fill_type='solid')
    legend_ws.cell(row=9, column=1).fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    legend_ws.cell(row=10, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    legend_ws.cell(row=11, column=1).fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    
    legend_ws.cell(row=14, column=1).fill = PatternFill(start_color='00AA00', end_color='00AA00', fill_type='solid')
    legend_ws.cell(row=15, column=1).fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    legend_ws.cell(row=16, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    legend_ws.cell(row=17, column=1).fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    
    # Format legend headers and sections
    legend_ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    legend_ws.cell(row=3, column=1).font = Font(bold=True, size=12, color='000080')
    legend_ws.cell(row=5, column=1).font = Font(bold=True, size=12, underline='single')
    legend_ws.cell(row=7, column=1).font = Font(bold=True, italic=True)
    legend_ws.cell(row=13, column=1).font = Font(bold=True, italic=True)
    legend_ws.cell(row=20, column=1).font = Font(bold=True, size=12, underline='single')
    legend_ws.cell(row=22, column=1).font = Font(bold=True, color='00AA00')
    
    # Find where "Lower is Better" starts and format it
    lower_is_better_row = 22 + len(HIGHER_IS_BETTER) + 2
    legend_ws.cell(row=lower_is_better_row, column=1).font = Font(bold=True, color='FF0000')
    
    # Highlight pred_f1 in the legend
    for row in range(22, 22 + len(HIGHER_IS_BETTER) + 1):
        cell_value = legend_ws.cell(row=row, column=2).value
        if cell_value and PRIMARY_METRIC in cell_value:
            legend_ws.cell(row=row, column=2).font = Font(bold=True, color='000080')
    
    # Add borders to color samples
    for row in [8, 9, 10, 11, 14, 15, 16, 17]:
        legend_ws.cell(row=row, column=1).border = thin_border
    
    # Adjust legend column widths
    legend_ws.column_dimensions['A'].width = 25
    legend_ws.column_dimensions['B'].width = 40
    
    # Add descriptions sheet
    desc_ws = wb.create_sheet("Metric Descriptions")
    
    # Headers for descriptions
    desc_headers = ["Metric", "Description (English)", "Description (中文)"]
    for col_idx, header in enumerate(desc_headers, start=1):
        cell = desc_ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add descriptions for each column
    row_idx = 2
    for col_name in df.columns:
        if col_name in COLUMN_DESCRIPTIONS:
            desc_ws.cell(row=row_idx, column=1, value=col_name)
            desc_ws.cell(row=row_idx, column=2, value=COLUMN_DESCRIPTIONS[col_name]['en'])
            desc_ws.cell(row=row_idx, column=3, value=COLUMN_DESCRIPTIONS[col_name]['zh'])
            
            # Apply appropriate font color based on metric type
            if col_name == PRIMARY_METRIC:
                desc_ws.cell(row=row_idx, column=1).font = Font(color='000080', bold=True, size=12)
                desc_ws.cell(row=row_idx, column=1).fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
                desc_ws.cell(row=row_idx, column=2).fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
                desc_ws.cell(row=row_idx, column=3).fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
                # Apply thick border to primary metric row
                for col in range(1, 4):
                    desc_ws.cell(row=row_idx, column=col).border = thick_border
            elif col_name in HIGHER_IS_BETTER:
                desc_ws.cell(row=row_idx, column=1).font = Font(color='00AA00', bold=True)
                for col in range(1, 4):
                    desc_ws.cell(row=row_idx, column=col).border = thin_border
            elif col_name in LOWER_IS_BETTER:
                desc_ws.cell(row=row_idx, column=1).font = Font(color='FF0000', bold=True)
                for col in range(1, 4):
                    desc_ws.cell(row=row_idx, column=col).border = thin_border
            else:
                desc_ws.cell(row=row_idx, column=1).font = Font(bold=True)
                for col in range(1, 4):
                    desc_ws.cell(row=row_idx, column=col).border = thin_border
            
            row_idx += 1
    
    # Adjust alignment for description cells
    for row in desc_ws.iter_rows(min_row=2, max_row=row_idx-1):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    desc_ws.column_dimensions['A'].width = 30
    desc_ws.column_dimensions['B'].width = 60
    desc_ws.column_dimensions['C'].width = 50
    
    # Set row heights for better readability
    for row in range(2, row_idx):
        desc_ws.row_dimensions[row].height = 45
    
    # Save the workbook
    wb.save(output_path)