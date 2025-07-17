#!/usr/bin/env python3
"""
Script to reorganize dataset files into a cleaner structure using song names.
"""

import os
import shutil
from pathlib import Path
from openpyxl import load_workbook
from typing import NamedTuple

class Row(NamedTuple):
    id_: str
    song_name: str
    singer: str
    hash_: str
    is_original: bool

def sanitize_filename(filename):
    """Sanitize filename by removing/replacing problematic characters."""
    # Replace problematic characters with underscores
    problematic_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
    for char in problematic_chars:
        filename = filename.replace(char, '_')
    return filename.strip()

def reorganize_dataset():
    # Read the mapping table
    base_path = Path("piano-song_test")
    xlsx_file = base_path / "钢琴曲+原唱_random20+20.xlsx"
    
    print("Reading mapping table...")
    workbook = load_workbook(xlsx_file)
    sheet = workbook["Sheet1"]
    
    table: list[Row] = []
    for row in sheet.iter_rows(min_row=2, max_col=10, values_only=True):
        if row[0] is None:
            break
        if row[4] == 0:
            table.append(Row(*[str(i) for i in row[:5]]))
    
    print(f"Found {len(table)} songs in the mapping table")
    
    # Create the new dataset directory structure
    dataset_dir = Path("dataset_dirty")
    dataset_dir.mkdir(exist_ok=True)
    
    # Process each song
    for row in table:
        song_name = sanitize_filename(row.song_name + "(" + row.singer + ")")
        song_dir = dataset_dir / song_name
        song_dir.mkdir(exist_ok=True)
        
        print(f"Processing: {row.song_name}")
        
        # Define source and destination paths
        file_mappings = [
            # (source_path, destination_filename)
            (base_path / "song_note" / f"{row.hash_}.mid", f"{song_name}.m.mid"),
            (base_path / "piano" / row.id_ / f"{row.id_}_transcription.mid", f"{song_name}.t.mid"),
            (base_path / "piano" / row.id_ / f"{row.id_}_rectified.mid", f"{song_name}.gt.mid"),
            (base_path / "piano" / row.id_ / f"{row.id_}.mid", f"{song_name}.bl.mid"),
        ]
        
        # Check for additional files (mp3, opus)
        additional_files = []
        audio_dir = base_path / "piano_melody_test_samples" / row.id_
        for ext in ['.mp3', '.opus']:
            audio_file = audio_dir / f"{row.id_}{ext}"
            if audio_file.exists():
                additional_files.append((audio_file, f"{song_name}{ext}"))
        
        # Copy files to the new structure
        for source_path, dest_filename in file_mappings + additional_files:
            dest_path = song_dir / dest_filename
            if source_path.exists():
                # print(f"  Copying: {source_path} -> {dest_path}")
                shutil.copy2(source_path, dest_path)
            else:
                print(f"  WARNING: Source file not found: {source_path}")
    
    print(f"\nDataset reorganization complete!")
    print(f"New structure created in: {dataset_dir.absolute()}")
    print(f"Total songs processed: {len(table)}")
    
    # Create a summary file
    summary_file = dataset_dir / "dataset_summary.txt"
    with open(summary_file, 'w', encoding='utf-8') as f:
        f.write("Dataset Summary\n")
        f.write("===============\n\n")
        f.write(f"Total songs: {len(table)}\n\n")
        f.write("Songs included:\n")
        for row in table:
            f.write(f"- {row.song_name} (by {row.singer})\n")
        f.write("\nFile naming convention:\n")
        f.write("- {song_name}.m.mid     - Melody annotations\n")
        f.write("- {song_name}.x.mid   - Transcribed MIDI\n")
        f.write("- {song_name}.gt.mid       - Ground truth MIDI\n")
        f.write("- {song_name}.bl.mid        - Baseline MIDI\n")
        f.write("- {song_name}.{ext}         - Audio files (if available)\n")
    
    print(f"Summary saved to: {summary_file}")

if __name__ == "__main__":
    reorganize_dataset() 